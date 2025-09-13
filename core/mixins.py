"""
Mixin per funzionalità avanzate admin - CerCollettiva
Export/Import Excel, Filtri avanzati, Report personalizzati
"""

# import io
import logging
from datetime import datetime

from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db import models
from django.http import HttpResponse
from django.utils import timezone

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class ExcelExportMixin:
    """
    Mixin per export Excel di modelli Django
    """

    def export_to_excel(self, request, queryset):
        """
        Export queryset to Excel file
        """
        try:
            # Crea workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"{self.model._meta.verbose_name_plural}"

            # Prepara dati
            data = []
            headers = []

            # Ottieni campi del modello
            fields = self.get_export_fields()

            # Aggiungi headers
            for field_name, field_label in fields.items():
                headers.append(field_label)

            # Aggiungi dati
            for obj in queryset:
                row = []
                for field_name, field_label in fields.items():
                    value = self.get_field_value(obj, field_name)
                    row.append(value)
                data.append(row)

            # Crea DataFrame
            df = pd.DataFrame(data, columns=headers)

            # Aggiungi al worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            # Formatta headers
            self._format_excel_headers(ws, len(headers))

            # Auto-adjust column widths
            self._auto_adjust_columns(ws)

            # Crea response
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            )
            filename = (
                f"{self.model._meta.verbose_name_plural}_"
                f"{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

            # Salva workbook
            wb.save(response)

            # Log success
            logger.info(
                f"Excel export completed for {self.model._meta.label}: "
                f"{queryset.count()} records"
            )

            return response

        except Exception as e:
            logger.error(f"Excel export failed for {self.model._meta.label}: {str(e)}")
            messages.error(request, f"Errore durante l'export: {str(e)}")
            return None

    def get_export_fields(self):
        """
        Ottieni campi per export - da sovrascrivere nelle classi figlie
        """
        fields = {}
        for field in self.model._meta.fields:
            if not field.name.startswith("_"):
                fields[field.name] = field.verbose_name or field.name
        return fields

    def get_field_value(self, obj, field_name):
        """
        Ottieni valore del campo per export
        """
        try:
            # Gestisci relazioni
            if "__" in field_name:
                parts = field_name.split("__")
                value = obj
                for part in parts:
                    if hasattr(value, part):
                        value = getattr(value, part)
                    else:
                        return ""
                return str(value) if value else ""

            # Campo normale
            value = getattr(obj, field_name)

            # Gestisci tipi speciali
            if isinstance(value, models.Model):
                return str(value)
            elif isinstance(value, datetime):
                return value.strftime("%Y-%m-%d %H:%M:%S")
            elif value is None:
                return ""
            else:
                return str(value)

        except Exception as e:
            logger.warning(f"Error getting field value for {field_name}: {str(e)}")
            return ""

    def _format_excel_headers(self, ws, num_columns):
        """
        Formatta headers Excel
        """
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Applica stile a headers
        for col in range(1, num_columns + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    def _auto_adjust_columns(self, ws):
        """
        Auto-adjust column widths
        """
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Max width 50
            ws.column_dimensions[column_letter].width = adjusted_width


class ExcelImportMixin:
    """
    Mixin per import Excel di modelli Django
    """

    def import_from_excel(self, request, file):
        """
        Import data from Excel file
        """
        try:
            # Leggi Excel
            df = pd.read_excel(file)

            # Valida struttura
            if not self._validate_excel_structure(df):
                messages.error(request, "Struttura file Excel non valida")
                return False

            # Processa dati
            success_count = 0
            error_count = 0
            errors = []

            for index, row in df.iterrows():
                try:
                    obj = self._create_object_from_row(row)
                    if obj:
                        success_count += 1
                except Exception as e:
                    error_count += 1
                    errors.append(f"Riga {index + 2}: {str(e)}")

            # Mostra risultati
            if success_count > 0:
                messages.success(
                    request, f"Import completato: {success_count} record importati"
                )

            if error_count > 0:
                messages.warning(
                    request,
                    f"Import completato con errori: {error_count} record non importati",
                )
                for error in errors[:5]:  # Mostra solo primi 5 errori
                    messages.error(request, error)

            # Log results
            logger.info(
                f"Excel import completed for {self.model._meta.label}: "
                f"{success_count} success, {error_count} errors"
            )

            return True

        except Exception as e:
            logger.error(f"Excel import failed for {self.model._meta.label}: {str(e)}")
            messages.error(request, f"Errore durante l'import: {str(e)}")
            return False

    def _validate_excel_structure(self, df):
        """
        Valida struttura file Excel
        """
        required_fields = self.get_required_import_fields()

        for field in required_fields:
            if field not in df.columns:
                return False

        return True

    def get_required_import_fields(self):
        """
        Ottieni campi obbligatori per import - da sovrascrivere
        """
        return []

    def _create_object_from_row(self, row):
        """
        Crea oggetto dal row Excel
        """
        try:
            # Prepara dati
            data = {}
            for field_name, value in row.items():
                if pd.notna(value):
                    data[field_name] = value

            # Crea oggetto
            obj = self.model(**data)
            obj.full_clean()  # Valida
            obj.save()

            return obj

        except ValidationError as e:
            raise Exception(f"Errore validazione: {str(e)}")
        except Exception as e:
            raise Exception(f"Errore creazione oggetto: {str(e)}")


class AdvancedFilterMixin:
    """
    Mixin per filtri avanzati
    """

    def get_advanced_filters(self, request):
        """
        Ottieni filtri avanzati
        """
        filters = {}

        # Filtro per data
        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")

        if date_from:
            filters["created_at__gte"] = date_from
        if date_to:
            filters["created_at__lte"] = date_to

        # Filtro per status
        status = request.GET.get("status")
        if status:
            filters["is_active"] = status == "active"

        # Filtro per tipo
        obj_type = request.GET.get("type")
        if obj_type:
            filters["type"] = obj_type

        return filters

    def apply_advanced_filters(self, queryset, request):
        """
        Applica filtri avanzati al queryset
        """
        filters = self.get_advanced_filters(request)
        return queryset.filter(**filters)


class ReportMixin:
    """
    Mixin per report personalizzati
    """

    def generate_report(self, request, queryset, report_type="summary"):
        """
        Genera report personalizzato
        """
        try:
            if report_type == "summary":
                return self._generate_summary_report(queryset)
            elif report_type == "detailed":
                return self._generate_detailed_report(queryset)
            elif report_type == "analytics":
                return self._generate_analytics_report(queryset)
            else:
                raise ValueError(f"Tipo report non supportato: {report_type}")

        except Exception as e:
            logger.error(
                f"Report generation failed for {self.model._meta.label}: {str(e)}"
            )
            messages.error(request, f"Errore generazione report: {str(e)}")
            return None

    def _generate_summary_report(self, queryset):
        """
        Genera report riassuntivo
        """
        return {
            "total_count": queryset.count(),
            "active_count": queryset.filter(is_active=True).count(),
            "inactive_count": queryset.filter(is_active=False).count(),
            "created_today": queryset.filter(
                created_at__date=timezone.now().date()
            ).count(),
        }

    def _generate_detailed_report(self, queryset):
        """
        Genera report dettagliato
        """
        return {
            "data": list(queryset.values()),
            "summary": self._generate_summary_report(queryset),
        }

    def _generate_analytics_report(self, queryset):
        """
        Genera report analytics
        """
        return {
            "summary": self._generate_summary_report(queryset),
            "trends": self._get_trends(queryset),
            "statistics": self._get_statistics(queryset),
        }

    def _get_trends(self, queryset):
        """
        Ottieni trend per report
        """
        # Implementazione base - da estendere
        return {}

    def _get_statistics(self, queryset):
        """
        Ottieni statistiche per report
        """
        # Implementazione base - da estendere
        return {}
